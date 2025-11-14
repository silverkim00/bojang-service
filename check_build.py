import os
import sys
import importlib
import pathlib
import hashlib
import platform

print("PYTHON", platform.python_version())
print("CWD", os.getcwd())
print("PYTHONPATH", sys.path)

MODULES = (
    "pdf_xlsx.coverage_processor",
    "pdf_xlsx.pdf_processor",
    "pdf_xlsx.excel_handler",
    "pdf_xlsx.mapping_rules",
    "pdf_xlsx.utils",
)

def import_ok(mod_name: str):
    try:
        m = importlib.import_module(mod_name)
        p = getattr(m, "__file__", None)
        print("IMPORT_OK", mod_name, p)
        try:
            if p and pathlib.Path(p).is_file():
                st = pathlib.Path(p).stat()
                if st.st_size <= 1_000_000:
                    md5 = hashlib.md5(pathlib.Path(p).read_bytes()).hexdigest()
                    print("MD5", mod_name, md5)
        except Exception as e:
            print("MD5_SKIP", mod_name, type(e).__name__, e)
        return True
    except Exception as e:
        print("IMPORT_FAIL", mod_name, type(e).__name__, e)
        return False

all_ok = True
for name in MODULES:
    all_ok &= import_ok(name)

if not all_ok:
    sys.exit("Some project modules failed to import.")

# ===== 템플릿 존재/열기 확인 (A3/A4 모두) =====
from openpyxl import load_workbook

env_single = os.environ.get("TEMPLATE_FILE")  # 하위호환(단일)
env_a3 = os.environ.get("TEMPLATE_FILE_A3", "/app/templates/base_template.xlsx")
env_a4 = os.environ.get("TEMPLATE_FILE_A4", "/app/templates/base_template2.xlsx")

candidates = []
if env_single:
    candidates.append(("ENV_SINGLE", pathlib.Path(env_single)))
candidates.append(("A3", pathlib.Path(env_a3)))
candidates.append(("A4", pathlib.Path(env_a4)))

seen = {}
for key, p in candidates:
    if not p.exists():
        print("TPL_MISSING", key, str(p))
        continue
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        print("TPL_OK", key, str(p), wb.sheetnames[:1])
        try:
            size = p.stat().st_size
            seen[key] = (str(p), size)
        except Exception:
            seen[key] = (str(p), None)
        wb.close()
    except Exception as e:
        raise SystemExit(f"openpyxl failed ({key}): {p} -> {e}")

if "A3" not in seen or "A4" not in seen:
    raise SystemExit("Both A3 and A4 templates must exist in image.")

print(f"[OK] templates: A3={seen['A3'][0]} ({seen['A3'][1]} bytes), "
      f"A4={seen['A4'][0]} ({seen['A4'][1]} bytes)")

# ===== Django 설정 모듈 점검 =====
DJANGO_SETTINGS_MODULE = os.environ.get("DJANGO_SETTINGS_MODULE", "bojang_api.settings")
print("DJANGO_SETTINGS_MODULE", DJANGO_SETTINGS_MODULE)

try:
    importlib.import_module(DJANGO_SETTINGS_MODULE)
    print("DJANGO_SETTINGS_IMPORT_OK")
except Exception as e:
    raise SystemExit(f"DJANGO_SETTINGS_IMPORT_FAIL: {e}")

try:
    import django
    django.setup()
    print("DJANGO_SETUP_OK")
except Exception as e:
    print("DJANGO_SETUP_WARN", type(e).__name__, e)

print("CHECK_BUILD_OK")
