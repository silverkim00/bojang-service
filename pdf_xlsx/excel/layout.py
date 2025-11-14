# -*- coding: utf-8 -*-
"""
excel_layout.py — v2025-11-06
역할: 엑셀 레이아웃/서식 전담(헤더 페인트, 열 복제, 줄바꿈, 너비/행높이)
주의: excel_handler.py v2025-11-05의 동일 동작 유지(이관본).
"""
from __future__ import annotations
import re
from copy import copy
from typing import List, Optional, Set
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from .. import config
from ..logger_setup import logger
# -----------------------------
# 공용 색상/헤더 컬러
# -----------------------------
GREEN = "006100"   # 갱신(초록)
BLACK = "000000"
RED   = "C00000"   # 14행(기타) 빨강

HEADER_FALLBACK_RGB = "1F3B57"   # 네이비
HEADER_TEXT_RGB     = "FFFF00"   # 노란 글자

# -----------------------------
# 범용 유틸/범위
# -----------------------------
def _col_letter(idx: int) -> str:
    return get_column_letter(idx)

COVERAGE_START_ROW = 16
try:
    COVERAGE_END_ROW = max(config.HARDCODED_ROW_MAP.values())
except Exception:
    COVERAGE_END_ROW = 100

# 실손 예외행(개행/도색 예외, 템플릿 스타일 유지)
_SILSON_NAMES = [
    "질병,상해 입원의료비",
    "질병,상해 통원의료비",
    "도수,체외충격파,증식",
    "비급여주사료",
    "비급여영상진단MRI",
]
_SILSON_ROWS: Set[int] = set(
    int(config.HARDCODED_ROW_MAP[nm]) for nm in _SILSON_NAMES if nm in config.HARDCODED_ROW_MAP
)

# 자동 줄바꿈 대상(실손 제외)
_WRAP_ROWS: Set[int] = set(r for r in range(COVERAGE_START_ROW, COVERAGE_END_ROW + 1) if r not in _SILSON_ROWS)

def _apply_fill(cell, rgb: Optional[str]):
    rgb_hex = (rgb or "").upper() or HEADER_FALLBACK_RGB
    cell.fill = PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")

# -----------------------------
# 1) 헤더 페인트 & 열 복제
# -----------------------------
def _paint_header_cell(ws: Worksheet, col: int):
    src = ws.cell(row=1, column=7)  # G1 기준
    dst = ws.cell(row=1, column=col)
    dst.value = None  # 헤더 값 초기화만 허용

    rgb = None
    if getattr(src.fill, "fill_type", None) == "solid":
        rgb = (src.fill.start_color.rgb or src.fill.fgColor.rgb)
        if rgb:
            rgb = rgb[-6:]
    _apply_fill(dst, rgb or HEADER_FALLBACK_RGB)

    dst.font = Font(name=(src.font.name if src.font else None),
                    size=(src.font.size if src.font else None),
                    b=True, color=HEADER_TEXT_RGB)
    dst.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    dst.border = copy(src.border)
    dst.number_format = src.number_format

def _clone_column_layout(ws: Worksheet, src_col: int, dst_col: int, max_row: Optional[int] = None) -> None:
    """
    템플릿 서식을 대상 열로 복제한다. **데이터 값은 보존**한다.
    - d.value 를 건드리지 않는다(값 소실 방지).
    - 헤더(1행)는 별도 처리.
    """
    if max_row is None:
        max_row = ws.max_row
    src_col = 7  # G열 고정(템플릿 기준열)
    s_col = _col_letter(src_col)
    d_col = _col_letter(dst_col)

    # 대상 열에 걸친 병합 해제(1행 제외)
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col <= dst_col <= rng.max_col:
            if rng.min_row == 1 and rng.max_row == 1:
                continue
            to_unmerge.append(rng)
    for rng in to_unmerge:
        ws.unmerge_cells(str(rng))

    # 폭 복제
    if ws.column_dimensions.get(s_col) and ws.column_dimensions.get(s_col).width:
        ws.column_dimensions[d_col].width = ws.column_dimensions[s_col].width

    # 1행 헤더 복제(헤더만 값 초기화 허용)
    _paint_header_cell(ws, dst_col)

    # 2행~max_row 레이아웃 복제(데이터 값은 유지)
    for r in range(2, max_row + 1):
        s = ws.cell(row=r, column=src_col)
        d = ws.cell(row=r, column=dst_col)
        if isinstance(d, MergedCell):
            continue

        # d.value = None  # **금지**: 값 소실 방지
        d.border        = copy(s.border)
        d.number_format = s.number_format
        d.protection    = copy(s.protection)

        # 정렬: 담보 구간은 wrap_text 기본 ON
        base_align = s.alignment or Alignment()
        want_wrap = (r in _WRAP_ROWS)
        d.alignment = Alignment(
            horizontal = base_align.horizontal or "left",
            vertical   = base_align.vertical or "top",
            wrap_text  = True if want_wrap else (getattr(base_align, "wrap_text", False))
        )

        # Fill 복제(실손/비담보 구간만)
        if r in _SILSON_ROWS:
            d.fill = copy(s.fill)
        elif not (COVERAGE_START_ROW <= r <= COVERAGE_END_ROW):
            d.fill = copy(s.fill)

        # Font 복제(담보/헤더는 검정 초기화) — 값은 유지
        if r in _SILSON_ROWS:
            d.font = copy(s.font) if s.font else Font()
        elif r <= 15 or (COVERAGE_START_ROW <= r <= COVERAGE_END_ROW):
            f = s.font or Font()
            d.font = Font(
                name=f.name, size=f.size, b=f.b, i=f.i,
                underline=f.underline, strike=f.strike, color=BLACK
            )
        else:
            d.font = copy(s.font) if s.font else Font(color=BLACK)

def _extend_header_band(ws: Worksheet, last_col: int) -> None:
    try:
        for c in range(7, last_col + 1):
            _paint_header_cell(ws, c)
        logger.info("[EXCEL] header: painted row 1 cells up to %s (no merge).", _col_letter(last_col))
    except Exception:
        logger.info("[EXCEL] header: paint skipped (non-critical)")

# -----------------------------
# 2) 자동 폭/행 높이 + 담보 줄바꿈
# -----------------------------
def _auto_adjust_col_widths(ws: Worksheet, start_col: int, end_col: int) -> None:
    logger.info("[EXCEL] auto-width: start=%s, end=%s", start_col, end_col)
    for c in range(start_col, end_col + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).replace("\n", " ").strip()
            if not s:
                continue
            max_len = max(max_len, len(s))
        width = 14.0
        if max_len > 12:
            width = min(26.0, 10.0 + max_len * 0.7)
        ws.column_dimensions[_col_letter(c)].width = round(width, 1)
        logger.info("[EXCEL] auto-width: set col=%s width=%.1f", _col_letter(c), width)

COMMA_TOPLEVEL_RX = re.compile(r",(?![^()]*\))")
LONG_TOKEN_CH = 18

def _split_by_top_level_comma(text: str) -> List[str]:
    if not isinstance(text, str): return []
    parts = [p.strip() for p in COMMA_TOPLEVEL_RX.split(text)]
    return [p for p in parts if p]

def _is_long_token(tok: str) -> bool:
    return len((tok or "").replace(" ", "")) >= LONG_TOKEN_CH

def _apply_coverage_linebreaks(row_idx: int, text: str) -> str:
    if not isinstance(text, str): return text
    if not (COVERAGE_START_ROW <= row_idx <= COVERAGE_END_ROW) or (row_idx in _SILSON_ROWS): return text
    if "\n" in text: return text
    parts = _split_by_top_level_comma(text)
    if len(parts) <= 2: return text
    lines: List[str] = []
    i, n = 0, len(parts)
    while i < n:
        group = parts[i:i+2]; i += 2
        if any(_is_long_token(t) for t in group):
            lines.extend(group)
        else:
            lines.append(", ".join(group))
    return ",\n".join(lines)

def _finalize_coverage_layout(ws: Worksheet, start_col: int, end_col: int) -> None:
    try:
        base_h = ws.row_dimensions[COVERAGE_START_ROW].height or 16.5
    except Exception:
        base_h = 16.5
    if base_h > 22.0: base_h = 16.5

    cap = 3.5
    for r in range(COVERAGE_START_ROW, COVERAGE_END_ROW + 1):
        # 실손 구간: 개행 금지, 기본 높이 유지
        if r in _SILSON_ROWS:
            if ws.row_dimensions[r].height is None:
                ws.row_dimensions[r].height = base_h
            continue

        max_lines = 1
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            ba = cell.alignment or Alignment()
            if not getattr(ba, "wrap_text", False):
                cell.alignment = Alignment(horizontal=ba.horizontal or "left",
                                           vertical=ba.vertical or "top",
                                           wrap_text=True)
            v = cell.value
            if isinstance(v, str):
                nv = _apply_coverage_linebreaks(r, v)
                nv = re.sub(r"\s*,\s*\n\s*", ",\n", nv.strip())
                if nv != v: cell.value = nv
                lines = nv.count("\n") + 1
                flat = nv.replace("\n", " ")
                if len(flat) >= 90: lines = max(lines, 2)  # 매우 긴 단일 라인 보정
                max_lines = max(max_lines, lines)
        h = min(cap, max_lines) * (base_h * 1.10)
        ws.row_dimensions[r].height = h
