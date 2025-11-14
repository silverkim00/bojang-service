# -*- coding: utf-8 -*-
"""
excel_handler.py — v2025-11-06c (폴더 이관 완료)
- excel/ 패키지로 모듈 분리: layout/renewal/product/misc
"""
from __future__ import annotations

import re, shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Set

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ⬇ 패키지 내부 상대 임포트
from . import config
from .utils import _nz
from .logger_setup import logger
from .coverage_processor import process_coverages
from .mapping_rules import find_target_row

from .excel.layout import (
    _col_letter, GREEN, BLACK, RED,
    COVERAGE_START_ROW, COVERAGE_END_ROW, _SILSON_ROWS,
    _apply_coverage_linebreaks, _auto_adjust_col_widths,
    _finalize_coverage_layout, _clone_column_layout, _extend_header_band,
)
from .excel.renewal import _apply_richtext_tokens, _row_has_mixed_renewal, _RT_SUPPORTED
from .excel.product import (
    normalize_product_name, merge_products, is_renewal_product, display_product_name
)
from .excel.misc import build_row14_text, normalize_log_entries


# 갱신 훅(coverage 재판정에서 사용)
try:
    from .rules.ca import cov_is_renewal as cov_is_renewal
except Exception:
    def cov_is_renewal(cov: Dict) -> bool:
        s = f"{_nz(cov.get('name'))}|{_nz(cov.get('association_name'))}"
        if re.search(r"(비\s*갱신형|비갱신)", s, re.I):
            return False
        return ("갱신" in s)

try:
    from .rules.ca import _RE_NOT_RENEWAL as _CA_RE_NOT_RENEWAL
except Exception:
    _CA_RE_NOT_RENEWAL = re.compile(r"(비\s*갱신형|비갱신)", re.I)


def _trace(tag: str, msg: str) -> None:
    try:
        logger.info(f"TRACE[{tag}] {msg}")
    except Exception:
        pass


def create_analysis_report(
    data: dict,
    output_path: str,
    base_template_path: Optional[str] = None,
) -> Tuple[List[str], List[str]]:
    """
    템플릿 복제 → 열 확장/레이아웃 → 상품별 기입 → 갱신/특수행 도색 → 자동 폭/행높이 → 저장
    반환: (unmapped_log, excluded_log)
    """
    # --- 템플릿 복제 ---
    template_file = base_template_path or getattr(config, "TEMPLATE_FILE", None)
    if not template_file:
        logger.error("템플릿 파일 경로가 설정되지 않았습니다.(TEMPLATE_FILE / base_template_path)")
        return [], []

    try:
        shutil.copy(template_file, output_path)
    except FileNotFoundError:
        logger.error(f"템플릿 파일 '{template_file}'을 찾을 수 없습니다!")
        return [], []
    except PermissionError:
        logger.error(f"'{output_path}' 파일이 열려있어 저장할 수 없습니다. 파일을 닫고 다시 실행해주세요.")
        return [], []

    wb = load_workbook(output_path)
    ws = wb.active

    # --- 상품 병합/총계 ---
    products = merge_products(data.get("products") or [])
    total_product_count = len(products)

    total_monthly_premium = 0
    for p in products:
        s = (p.get("monthly_premium", "") or "").replace(",", "")
        try:
            total_monthly_premium += int(s)
        except Exception:
            pass

    ws["A1"].value = f"{data.get('customer_name','-')}님 보장내용 ( 총 {total_product_count}건 월 {total_monthly_premium:,}원 )"

    all_products_data: List[Dict[str, object]] = []
    full_unmapped_log, full_excluded_log = [], []

    logger.info("[EXCEL] using excel_handler v2025-11-06c")

    # --- 각 상품 처리 ---
    for product in products:
        raw_name = product.get("product_name", "") or ""

        flat: Dict[str, object] = {
            "회사명": product.get("company", ""),
            "상품명": normalize_product_name(raw_name),  # 숫자 컷 없음
            "가입일": product.get("contract_date", "정보없음"),
            "만기": product.get("maturity_period", ""),
            "납입기간": product.get("payment_period", ""),
        }

        # 월 보험료/납입 집계(있을 때만)
        s = (product.get("monthly_premium", "") or "").replace(",", "")
        if "납입완료" in s:
            flat["월 보험료"] = "납입완료"
        elif "보험료미제공" in s:
            flat["월 보험료"] = "보험료미제공"
        else:
            try:
                flat["월 보험료"] = int(s)
                if flat.get("가입일") != "정보없음":
                    try:
                        now = datetime.now()
                        contract_dt = datetime.strptime(flat["가입일"], "%Y-%m-%d")
                        months_paid = ((now.year - contract_dt.year) * 12 + (now.month - contract_dt.month) + 1) if now >= contract_dt else 0
                        m = re.search(r"(\d+)", str(flat.get("납입기간", "")))
                        if m:
                            years = int(m.group(1))
                            total_pay = years * 12
                            total_premium = (flat["월 보험료"] or 0) * total_pay
                            if months_paid > total_pay:
                                months_paid = total_pay
                            paid_premium = (flat["월 보험료"] or 0) * months_paid
                            flat["납입횟수"] = months_paid
                            flat["납입한 보험료"] = paid_premium
                            flat["총 보험료"] = total_premium
                            flat["예상납입 보험료"] = total_premium - paid_premium
                    except Exception:
                        pass
            except Exception:
                flat["월 보험료"] = product.get("monthly_premium", "")

        # --- 매핑 호출 ---
        mapped, unmapped, excluded = process_coverages(product)
        flat.update(mapped)
        if unmapped:
            full_unmapped_log.extend(unmapped)
        if excluded:
            full_excluded_log.extend(excluded)

        flat["_RAW_PRODUCT_NAME"] = raw_name

        # --- 14행(기타) 면제/지원 텍스트 구성 ---
        row14_text = build_row14_text(product, mapped)
        if row14_text:
            flat["기타"] = (f"{flat.get('기타')} {row14_text}".strip() if flat.get("기타") else row14_text)

        # --- 갱신 메타 취득 ---
        renew_all = set(int(r) for r in (flat.get("_ROW_RENEWAL_ALL") or []))
        renew_any = set(int(r) for r in (flat.get("_ROW_RENEWAL_ANY") or []))

        # 행별 커버리지 바인딩
        row_to_covs: Dict[int, List[Dict]] = {}
        for cov in (product.get("coverages") or []):
            nm = _nz(cov.get("name")); assoc = _nz(cov.get("association_name"))
            tgt_label = find_target_row(nm, assoc)
            row = config.HARDCODED_ROW_MAP.get(tgt_label)
            if isinstance(row, int) and (COVERAGE_START_ROW <= row <= COVERAGE_END_ROW) and (row not in _SILSON_ROWS):
                row_to_covs.setdefault(int(row), []).append(cov)

        # renew_any 재산출(비갱신 우선)
        recomputed_any: Set[int] = set()
        for row, covs in row_to_covs.items():
            if not covs:
                continue
            has_nonrenewal = any(bool(_CA_RE_NOT_RENEWAL.search(f"{_nz(c.get('name'))}|{_nz(c.get('association_name'))}")) for c in covs)
            has_renewal    = any(cov_is_renewal(c) for c in covs)
            if has_nonrenewal:
                _trace("RENEWAL_DECIDE_ROW", f"row={row} nonrenewal=True renewal={has_renewal} -> drop"); continue
            if has_renewal:
                recomputed_any.add(row)
                _trace("RENEWAL_DECIDE_ROW", f"row={row} nonrenewal=False renewal=True -> add")
        renew_any = recomputed_any | {int(r) for r in renew_any if r not in recomputed_any and r not in row_to_covs}

        # 실제 기입될 담보행(실손 제외)
        mapped_rows = {
            int(row) for label, row in config.HARDCODED_ROW_MAP.items()
            if (label in flat) and isinstance(row, int)
            and (COVERAGE_START_ROW <= int(row) <= COVERAGE_END_ROW)
            and (int(row) not in _SILSON_ROWS)
        }

        # 가배책 ANY 강제
        _GABAECHEK_ROW = (
            config.HARDCODED_ROW_MAP.get("일상생활배상책임")
            or config.HARDCODED_ROW_MAP.get("가족/일상/자녀배상")
            or 56
        )
        if (_GABAECHEK_ROW >= COVERAGE_START_ROW and _GABAECHEK_ROW <= COVERAGE_END_ROW):
            if any(int(row) == _GABAECHEK_ROW for row in mapped_rows):
                renew_any = set(renew_any) | {_GABAECHEK_ROW}
                logger.info("[TRACE][RENEWAL_FORCE] row=%s label=일상생활배상책임", _GABAECHEK_ROW)

        # 전체 갱신 판단
        force_full_by_name = is_renewal_product(flat.get("_RAW_PRODUCT_NAME", ""))
        full_by_subset = (bool(mapped_rows) and mapped_rows.issubset(renew_all))
        flat["_ALL_COVS_RENEWAL"] = bool(force_full_by_name or full_by_subset)
        flat["_RENEWAL_ROWS_ANY"] = [] if flat["_ALL_COVS_RENEWAL"] else sorted(renew_any)
        flat["_ROW_TO_COVS"] = row_to_covs

        all_products_data.append(flat)

    # --- 열 확장/레이아웃 ---
    base_slots = 5   # C..G
    need_extra = max(0, len(all_products_data) - base_slots)
    if need_extra > 0:
        for _ in range(need_extra):
            ws.insert_cols(8)
            _clone_column_layout(ws, src_col=7, dst_col=8, max_row=ws.max_row)
    _extend_header_band(ws, last_col=ws.max_column)

    # --- 데이터 기록 ---
    start_col = 3  # C
    end_col = start_col + len(all_products_data) - 1
    logger.info("[EXCEL] write cols %s..%s", _col_letter(start_col), _col_letter(end_col))

    for col_idx, pdata in enumerate(all_products_data, start=start_col):
        if col_idx >= 8:
            _clone_column_layout(ws, src_col=7, dst_col=col_idx, max_row=ws.max_row)

        disp_name = display_product_name(
            pdata.get("_RAW_PRODUCT_NAME", ""),
            pdata.get("상품명", ""),
            pdata.get("_ALL_COVS_RENEWAL", False)
        )

        # 담보 컬럼 초기화(검정)
        for r in range(COVERAGE_START_ROW, COVERAGE_END_ROW + 1):
            if r in _SILSON_ROWS:
                continue
            cell = ws.cell(row=r, column=col_idx)
            f = cell.font or Font()
            cell.font = Font(name=f.name, size=f.size, b=False, color=BLACK)

        row_to_covs = pdata.get("_ROW_TO_COVS") or {}
        rt_rows_applied: Set[int] = set()

        # 셀 값 기입
        for label, row_num in config.HARDCODED_ROW_MAP.items():
            r = int(row_num)
            cell = ws.cell(row=r, column=col_idx)

            if label == "상품명":
                cell.value = disp_name
                cell.font = Font(color=(GREEN if str(disp_name).startswith("(갱)") else BLACK), b=True)
                continue

            if label == "주요 담보":
                if not cell.value:
                    cell.value = "주요 담보"
                continue

            val = pdata.get(label)
            if val is None:
                continue

            # 숫자 서식
            if label in ["월 보험료", "납입한 보험료", "예상납입 보험료", "총 보험료"] and isinstance(val, (int, float)):
                cell.value = f"{val:,}"
                continue

            # 담보 구간은 wrap 강제
            if COVERAGE_START_ROW <= r <= COVERAGE_END_ROW:
                base_align = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    horizontal=base_align.horizontal or "left",
                    vertical=base_align.vertical or "top",
                    wrap_text=True
                )

            # 문자열: 담보 줄바꿈(실손 제외)
            if isinstance(val, str) and (COVERAGE_START_ROW <= r <= COVERAGE_END_ROW) and (r not in _SILSON_ROWS):
                import re as _re
                val = _apply_coverage_linebreaks(r, val)
                val = _re.sub(r"\s*,\s*\n\s*", ",\n", val.strip())

            cell.value = val

            # [TOKEN-RT] ANY 행 + 문자열(콤마 포함)일 때 토큰별 부분서식 시도
            try:
                if (
                    isinstance(val, str)
                    and (COVERAGE_START_ROW <= r <= COVERAGE_END_ROW)
                    and (r not in _SILSON_ROWS)
                    and (r in (pdata.get("_RENEWAL_ROWS_ANY") or []))
                    and not bool(pdata.get("_ALL_COVS_RENEWAL", False))
                ):
                    applied = _apply_richtext_tokens(cell, val, row_to_covs.get(r, []))
                    if applied:
                        rt_rows_applied.add(r)
            except Exception:
                pass

        # --- 색상 규칙(열별) ---
        if pdata.get("_ALL_COVS_RENEWAL", False):
            # 전체 갱신: 담보 전부 초록(실손 제외)
            for r in range(COVERAGE_START_ROW, COVERAGE_END_ROW + 1):
                if r in _SILSON_ROWS:
                    continue
                f = ws.cell(row=r, column=col_idx).font or Font()
                ws.cell(row=r, column=col_idx).font = Font(name=f.name, size=f.size, b=True, color=GREEN)
        else:
            # 일부 갱신: RT 미적용 행만 행 단위 초록
            for r in (pdata.get("_RENEWAL_ROWS_ANY") or []):
                if COVERAGE_START_ROW <= r <= COVERAGE_END_ROW and r not in _SILSON_ROWS:
                    if r not in rt_rows_applied:
                        if (not _RT_SUPPORTED) and _row_has_mixed_renewal((pdata.get("_ROW_TO_COVS") or {}).get(r, [])):
                            _trace("RENEWAL_ROW_FALLBACK_SKIP", f"row={r} mixed=True rt=False -> keep BLACK")
                            continue
                        f = ws.cell(row=r, column=col_idx).font or Font()
                        ws.cell(row=r, column=col_idx).font = Font(name=f.name, size=f.size, b=True, color=GREEN)

        # --- 14행(기타) 최종 빨강 강제(내용 있을 때만) ---
        r14 = int(config.HARDCODED_ROW_MAP.get("기타", 14))
        c14 = ws.cell(row=r14, column=col_idx)
        if isinstance(c14.value, str) and c14.value.strip():
            f = c14.font or Font()
            c14.font = Font(name=f.name, size=f.size, b=True, color=RED)

        # [GUARD] 쓰기 검증
        try:
            non_empty = 0
            for rr in range(COVERAGE_START_ROW, COVERAGE_END_ROW + 1):
                if rr in _SILSON_ROWS:
                    continue
                v = ws.cell(row=rr, column=col_idx).value
                if isinstance(v, str):
                    if v.strip():
                        non_empty += 1
                elif v is not None:
                    non_empty += 1
            _trace("EXCEL_WRITE_CHECK", f"col={_col_letter(col_idx)} non_empty={non_empty}")
        except Exception:
            pass

    # --- 자동 폭/행높이 ---
    _auto_adjust_col_widths(ws, start_col=3, end_col=ws.max_column)
    _finalize_coverage_layout(ws, start_col=3, end_col=ws.max_column)

    # --- 저장 ---
    try:
        wb.save(output_path)
        logger.info(f"-> '{Path(output_path).name}' 보고서 생성 완료.")
    except PermissionError:
        logger.error(f"'{output_path}' 파일이 열려있어 저장할 수 없습니다. 파일을 닫고 다시 실행해주세요.")

    return normalize_log_entries(full_unmapped_log), normalize_log_entries(full_excluded_log)
